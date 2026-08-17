#!/usr/bin/env python3
"""Generate zadanie-qa-kwalifikacje.xlsx with 3 sheets."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(size=10)
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
P1_FILL = PatternFill("solid", fgColor="F8CBAD")
P2_FILL = PatternFill("solid", fgColor="FFE699")
FAIL_FILL = PatternFill("solid", fgColor="F4B7B7")
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 32


def style_rows(ws, ncols, start_row, end_row, min_height=60):
    for r in range(start_row, end_row + 1):
        ws.row_dimensions[r].height = min_height
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.font = CELL_FONT
            cell.border = BORDER


# --- Sheet 1: Test cases ---
ws1 = wb.active
ws1.title = "Test cases"
tc_headers = [
    "ID",
    "Warunek wstępny",
    "Kroki",
    "Dane testowe",
    "Wynik oczekiwany",
    "Priorytet",
]
tc_rows = [
    (
        "TC-01",
        "Widget załadowany, cookies zaakceptowane",
        "1. Wpisz frazę w pole 'Nazwa kwalifikacji'\n2. Kliknij 'Szukaj'",
        "kucharz",
        "Liczba wyników = 11, lista kafelków renderuje się, chip 'Aktywne filtry' pokazuje 'kucharz', URL zawiera ?search=kucharz",
        "P1",
    ),
    (
        "TC-02",
        "j.w.",
        "1. Wpisz frazę bez ogonków\n2. Wyczyść pole, wpisz tę samą frazę z ogonkami\n3. Porównaj liczby wyników",
        "ksiegowy vs księgowy",
        "Obie frazy zwracają te same wyniki (fold-diakrytyków = standard PL). Faktycznie: 0 vs 568 → patrz BR-01",
        "P1",
    ),
    (
        "TC-03",
        "j.w.",
        "1. Wpisz frazę WIELKIMI literami\n2. Powtórz z mieszaną wielkością\n3. Porównaj z lowercase",
        "KSIĘGOWY, Księgowy, księgowy",
        "Case-insensitive — każdy wariant daje 568 wyników",
        "P2",
    ),
    (
        "TC-04",
        "j.w.",
        "1. Wpisz frazę ze spacjami wiodącymi i końcowymi\n2. Kliknij 'Szukaj'",
        "   księgowy   ",
        "Spacje są przycinane (trim), wynik: 568 wyników (jak baseline)",
        "P2",
    ),
    (
        "TC-05",
        "j.w.",
        "1. Wpisz frazę, która nie istnieje w rejestrze\n2. Odczekaj na wynik",
        "zzzxxxnonexistent12345",
        "Wyświetlony komunikat 'Brak danych', licznik 'Znaleziono: 0', brak crashu widgetu, chip aktywnego filtra pokazuje wpisaną frazę",
        "P1",
    ),
    (
        "TC-06",
        "j.w.",
        "1. Otwórz URL z parametrem ?search= w nowej karcie",
        "https://kwalifikacje.gov.pl/wyszukiwarka-kwalifikacji/?search=księgowy",
        "Widget uzupełnia pole 'Nazwa kwalifikacji' wartością z URL, wyniki są przefiltrowane (568), chip aktywnego filtra widoczny",
        "P1",
    ),
    (
        "TC-07",
        "j.w., wyniki > 15",
        "1. Załaduj widok z domyślnym limit=15\n2. Kliknij przycisk 'Więcej kwalifikacji'",
        "brak filtrów",
        "Doklejone kolejne 15 kafelków (razem 30), licznik 'Znaleziono: 20760' bez zmian, offset w URL rośnie do 15",
        "P2",
    ),
    (
        "TC-08",
        "j.w.",
        "1. Wpisz w URL wartość z <script> (XSS smoke)\n2. Powtórz z bardzo długim ciągiem\n3. Powtórz ze znakami specjalnymi",
        "<script>alert(1)</script>\naaaa…(2000 znaków)\n@#$%^&*()",
        "Widget renderuje treść jako tekst (escapowanie React), brak alertu, brak crashu, 'Znaleziono: 0', HTTP 200 z API",
        "P2",
    ),
]

ws1.append(tc_headers)
for row in tc_rows:
    ws1.append(row)

style_header(ws1, len(tc_headers))
style_rows(ws1, len(tc_headers), 2, len(tc_rows) + 1, min_height=90)

# Column widths
widths = [10, 30, 45, 30, 55, 12]
for i, w in enumerate(widths, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Priority coloring
for r in range(2, len(tc_rows) + 2):
    pcell = ws1.cell(row=r, column=6)
    if pcell.value == "P1":
        pcell.fill = P1_FILL
    elif pcell.value == "P2":
        pcell.fill = P2_FILL
    pcell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    pcell.font = Font(size=10, bold=True)

ws1.freeze_panes = "A2"


# --- Sheet 2: Raport z błędów ---
ws2 = wb.create_sheet("Raport z błędów")
bug_headers = [
    "ID",
    "Tytuł",
    "Środowisko",
    "Kroki reprodukcji",
    "Wynik oczekiwany",
    "Wynik faktyczny",
    "Severity",
    "Priority",
    "Uzasadnienie sev/pri",
    "Dowód",
]
bug_rows = [
    (
        "BR-01",
        "Wyszukiwarka i autocomplete zwracają 0 wyników dla fraz bez polskich znaków (np. 'ksiegowy' zamiast 'księgowy')",
        "Chromium 141 desktop, macOS 24.6.0, prod: kwalifikacje.gov.pl/wyszukiwarka-kwalifikacji/, widget qualifications-searcher (Shadow DOM), API zrk-api.ibe.edu.pl/pl/v1",
        "1. Otwórz https://kwalifikacje.gov.pl/wyszukiwarka-kwalifikacji/\n2. Zaakceptuj cookies\n3. W polu 'Nazwa kwalifikacji' wpisz 'ksiegowy' (bez ogonków), odczekaj na wynik\n4. Wyczyść pole, wpisz 'księgowy' (z ogonkami), porównaj\n5. Autocomplete: wpisz 'ksieg' (bez ogonków) — obserwuj listę podpowiedzi",
        "Obie frazy zwracają zbliżoną liczbę wyników (fold-diakrytyków = standard dla języka polskiego). Autocomplete pokazuje podpowiedzi 'księg…' dla obu wariantów wpisu.",
        "'ksiegowy' → Znaleziono: 0, komunikat 'Brak danych', 0 podpowiedzi autocomplete.\n'księgowy' → 568 wyników, 2+ podpowiedzi.\nAPI: GET zrk-api.ibe.edu.pl/pl/v1/qualifications?search=ksiegowy → count: 0.",
        "Major",
        "High",
        "Ogranicza dostępność funkcji core dla użytkowników bez PL klawiatury (mobilni, obcokrajowcy, kopiujący z zewnątrz). Portal rządowy → duża grupa użytkowników. Naprawa prosta po stronie API (unaccent w PostgreSQL).",
        "Screenshot: bug1-ksiegowy-no-results.png\nURL: ?search=ksiegowy (0) vs ?search=księgowy (568)\nAPI: curl 'https://zrk-api.ibe.edu.pl/pl/v1/qualifications/?search=ksiegowy' → {\"count\":0,...}",
    ),
    (
        "BR-02",
        "Pole filtra 'Szukaj dziedziny…' bez powiązanej etykiety — łamie WCAG 2.2 AA (1.3.1 i 4.1.2)",
        "Chromium 141 + macOS VoiceOver, prod: kwalifikacje.gov.pl/wyszukiwarka-kwalifikacji/, Shadow DOM widget",
        "1. Otwórz stronę wyszukiwarki\n2. Rozwiń sekcję 'Filtry' → 'Dziedzina' (lewa kolumna)\n3. Sfokusuj pole 'Szukaj dziedziny…' klawiszem Tab\n4. Włącz VoiceOver (⌘F5) — pole odczytywane jako 'edytowalne pole tekstowe' bez nazwy\n5. DevTools: w Shadow Root #qualifications-container znajdź input#:re: — sprawdź atrybuty",
        "Input ma widoczny <label for> lub aria-label opisujący rolę pola (np. 'Filtruj dziedziny na liście'). Wymaga tego WCAG 2.2 AA — 1.3.1 Info and Relationships oraz 4.1.2 Name Role Value.",
        "Input: aria-label=null, aria-labelledby=null, brak <label for>. Jedyny wskaźnik to placeholder='Szukaj dziedziny…'. Placeholder nie zastępuje etykiety.",
        "Minor",
        "Medium",
        "Funkcjonalnie działa — user widzi placeholder i może pisać. ALE łamie wymóg prawny WCAG 2.2 AA. Portal rządowy publikuje deklarację dostępności; ta niezgodność powinna być odnotowana lub naprawiona.",
        "DevTools: document.querySelector('#qualifications-container').shadowRoot.querySelector('input#:re:') → obiekt z aria-label=null, labels.length=0.\nPlaceholder wizualny: 'Szukaj dziedziny…'.\nPorównaj z prawidłowym #name-search — ten ma aria-label='Wyszukaj po nazwie…'",
    ),
]

ws2.append(bug_headers)
for row in bug_rows:
    ws2.append(row)

style_header(ws2, len(bug_headers))
style_rows(ws2, len(bug_headers), 2, len(bug_rows) + 1, min_height=180)

widths2 = [8, 40, 30, 45, 40, 40, 12, 12, 40, 40]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Severity/Priority coloring
sev_colors = {"Major": PatternFill("solid", fgColor="F4B7B7"), "Minor": PatternFill("solid", fgColor="FFE699")}
pri_colors = {"High": PatternFill("solid", fgColor="F4B7B7"), "Medium": PatternFill("solid", fgColor="FFE699")}

for r in range(2, len(bug_rows) + 2):
    sev = ws2.cell(row=r, column=7)
    pri = ws2.cell(row=r, column=8)
    if sev.value in sev_colors:
        sev.fill = sev_colors[sev.value]
    if pri.value in pri_colors:
        pri.fill = pri_colors[pri.value]
    for cell in (sev, pri):
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.font = Font(size=10, bold=True)

ws2.freeze_panes = "A2"


# --- Sheet 3: A11y checklist ---
ws3 = wb.create_sheet("A11y checklist")
a11y_headers = ["#", "Kryterium WCAG 2.2 AA", "Wynik", "Uzasadnienie"]
a11y_rows = [
    (
        "A11Y-1",
        "1.3.1 Info and Relationships + 4.1.2 Name, Role, Value — wszystkie kontrolki formularza mają programową nazwę i rolę",
        "FAIL",
        "Główne pole #name-search ma aria-label='Wyszukaj po nazwie…' ✓ i powiązany <label> 'Nazwa kwalifikacji' ✓. Pole #category-search ma <label> 'Kategoria kwalifikacji' ✓. ALE wewnętrzne pole filtra input#:re: ('Szukaj dziedziny…') nie ma żadnej etykiety — patrz BR-02. Wystarczy jedno pole bez etykiety, aby kryterium było niespełnione.",
    ),
    (
        "A11Y-2",
        "2.4.7 Focus Visible — element z fokusem klawiaturowym ma widoczny wskaźnik",
        "PASS",
        "Tabowanie po widgecie: outline 4px czerwony solid (getComputedStyle().outline === 'rgb(255, 0, 0) solid 4px') wyraźnie widoczny na kafelkach, przyciskach, polach input i checkboxach. Kontrast wystarczający na białym tle strony (Δ czerwony vs biały > 3:1).",
    ),
    (
        "A11Y-3",
        "2.1.1 Keyboard — cała funkcjonalność dostępna z klawiatury bez pułapek fokusa",
        "PASS",
        "Tab przechodzi kolejno: pole nazwy → pole kategorii → checkbox 'aktualnie można uzyskać' → przycisk 'Szukaj' → sekcja filtrów → kafelki wyników → 'Więcej kwalifikacji'. Autocomplete listbox otwiera się po 2+ znakach, opcje wybierane strzałkami + Enter, Escape zamyka listbox. Brak pułapek fokusa.",
    ),
]

ws3.append(a11y_headers)
for row in a11y_rows:
    ws3.append(row)

style_header(ws3, len(a11y_headers))
style_rows(ws3, len(a11y_headers), 2, len(a11y_rows) + 1, min_height=90)

widths3 = [10, 55, 12, 90]
for i, w in enumerate(widths3, start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w

for r in range(2, len(a11y_rows) + 2):
    wcell = ws3.cell(row=r, column=3)
    if wcell.value == "PASS":
        wcell.fill = PASS_FILL
    elif wcell.value == "FAIL":
        wcell.fill = FAIL_FILL
    wcell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    wcell.font = Font(size=10, bold=True)

ws3.freeze_panes = "A2"


# --- Sheet 4: Założenia ---
ws4 = wb.create_sheet("Założenia", 0)
ws4.append(["ZADANIE REKRUTACYJNE — TESTER MANUALNY (QA)"])
ws4.append(["Portal", "kwalifikacje.gov.pl"])
ws4.append(["Moduł", "Wyszukiwarka kwalifikacji"])
ws4.append(["URL", "https://kwalifikacje.gov.pl/wyszukiwarka-kwalifikacji/"])
ws4.append(["Data testu", "2026-07-08"])
ws4.append(["Autor", "Łukasz Posmyk"])
ws4.append([])
ws4.append(["ZAŁOŻENIA I DECYZJE INTERPRETACYJNE"])
assumptions = [
    ("Zakres testów", "Wyłącznie widget 'Wyszukiwarka kwalifikacji' (/wyszukiwarka-kwalifikacji/) — pole nazwy, pole kategorii, checkbox 'aktualnie można uzyskać', wyniki, brak wyników, deep-link ?search=, paginacja (offset, limit), filtry kategorii/dziedziny. Poza zakresem: wyszukiwarka instytucji, karty szczegółów, menu portalu, dialog cookies."),
    ("Środowisko", "macOS 24.6.0, Chromium 141 (Playwright) + Chrome 141 desktop. Widget: web-component #qualifications-container (React + MUI w Shadow DOM), ładowany z CDN components.ibe.edu.pl. API: https://zrk-api.ibe.edu.pl/pl/v1/qualifications/"),
    ("Definicja 'krytyczne'", "Przypadki blokujące/istotnie pogarszające podstawowy przepływ (search → wyniki) lub łamiące wymóg WCAG 2.2 AA (portal rządowy podlega ustawie o dostępności cyfrowej)."),
    ("Empty search", "?search= (pusty) zwraca wszystkie rekordy (20760) — interpretuję jako zamierzone (równoważne brakowi parametru), nie zgłaszam jako bug."),
    ("Różnica UI vs API", "UI = 20760, surowe API = 20761. Widget dodaje domyślny filtr is_visible=1 do zapytania — dlatego UI pokazuje 20760. Różnica wyjaśniona, nie zgłaszam jako bug."),
    ("Autocomplete", "Wyzwala się po 2+ znakach — traktuję jako standardowe zachowanie MUI Autocomplete."),
    ("Severity vs Priority", "Rozdzielone świadomie. Severity = techniczna dotkliwość (Blocker/Major/Minor/Trivial). Priority = biznesowa pilność (High/Medium/Low). Bug a11y może mieć niski Severity (funkcjonalnie działa) ale wysoki Priority (wymóg prawny)."),
]
for k, v in assumptions:
    ws4.append([k, v])

ws4.append([])
ws4.append(["ZAWARTOŚĆ SKOROSZYTU"])
ws4.append(["Zakładka", "Zawartość"])
ws4.append(["Test cases", "8 krytycznych scenariuszy (5 pozytywnych, 3 negatywne)"])
ws4.append(["Raport z błędów", "2 błędy: BR-01 (fold-diakrytyków), BR-02 (a11y label)"])
ws4.append(["A11y checklist", "3 punkty WCAG 2.2 AA z wynikami PASS/FAIL"])

ws4.column_dimensions["A"].width = 22
ws4.column_dimensions["B"].width = 110

ws4.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F4E79")
for row_idx in (8, 17):
    cell = ws4.cell(row=row_idx, column=1)
    cell.font = Font(bold=True, size=12, color="1F4E79")

# Wrap all data cells
for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row, max_col=2):
    for cell in row:
        cell.alignment = WRAP

for r in range(2, ws4.max_row + 1):
    ws4.row_dimensions[r].height = 40


output_path = "/Users/lukasz/Developer/PlayWrightAutomation/qa-task-kwalifikacje/zadanie-qa-kwalifikacje.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Sheets: {wb.sheetnames}")
