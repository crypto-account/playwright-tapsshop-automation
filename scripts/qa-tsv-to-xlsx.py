#!/usr/bin/env python3
"""
Convert a /qa-case run's two TSV files into a styled single-file XLSX.

Usage:
    scripts/qa-tsv-to-xlsx.py <base>
    scripts/qa-tsv-to-xlsx.py qa-runs/wyszukiwarka-instytucji-2026-08-13-run2

Reads:
    <base>-cases.tsv, <base>-report.tsv
    <base>-bugs.tsv   (optional — adds a third "Bug tracker" sheet)
    <base>.md         (optional — provides Feature / Priority / Date for the title bar)

Writes:
    <base>.xlsx       (2 or 3 sheets: "Test cases", "Execution report", optionally "Bug tracker")
"""

import csv, io, re, sys
from pathlib import Path
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.hyperlink import Hyperlink

# Title bar (row 1) — darker shade than header row
TITLE_FILL = PatternFill("solid", fgColor="305496")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_ALIGN = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=False)

# Header row (row 2)
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Calibri", size=10)
BODY_FONT_BOLD = Font(name="Calibri", size=10, bold=True)

# Columns whose body cells should render in bold (emphasises the case name)
BOLD_COLUMNS = {"Tytuł", "Title"}

# Category-colored columns that should stay in NORMAL weight (not bold).
# By default, colored cells get CATEGORY_FONT (bold) — this set overrides to plain body font.
CATEGORY_PLAIN_COLUMNS = {"Priorytet", "Priority"}

# Columns whose body cells are parsed for inline rich text:
#   "..." (double-quoted spans)  → bold
#   https?://… full URL          → hyperlink (blue + underline; whole cell links if only one URL)
#   /path?query relative path    → hyperlink expanded against the run's base URL
RICH_TEXT_COLUMNS = {
    "Kroki", "Steps",
    "Reprodukcja", "Repro",
    "Kroki reprodukcji",
    "Warunki wstępne", "Preconditions",
    "Dane testowe", "Test Data",
    "Oczekiwany rezultat", "Expected Result",
    "Rzeczywisty rezultat", "Actual Result",
    "Środowisko", "URL", "Wpływ na użytkownika",
}

# Rich-text fonts — InlineFont accepts colours as AARRGGBB
RT_PLAIN = InlineFont(rFont="Calibri", sz=10)
RT_BOLD  = InlineFont(rFont="Calibri", sz=10, b=True)
RT_LINK  = InlineFont(rFont="Calibri", sz=10, u="single", color="FF0563C1")
BODY_ALIGN_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
BODY_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

ZEBRA_FILL = PatternFill("solid", fgColor="F2F2F2")

THIN = Side(border_style="thin", color="D0D0D0")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

# Category fills — pastel backgrounds, dark text
PRIORITY_FILL = {
    "P1 · Krytyczny": PatternFill("solid", fgColor="FFCCCC"),
    "P2 · Wysoki":    PatternFill("solid", fgColor="FFF2CC"),
    "P3 · Niski":     PatternFill("solid", fgColor="D9D9D9"),
    # Legacy short codes still recognised
    "P1": PatternFill("solid", fgColor="FFCCCC"),
    "P2": PatternFill("solid", fgColor="FFF2CC"),
    "P3": PatternFill("solid", fgColor="D9D9D9"),
}
TYPE_FILL = {
    "positive": PatternFill("solid", fgColor="C6EFCE"),
    "negative": PatternFill("solid", fgColor="FFC7CE"),
    "edge":     PatternFill("solid", fgColor="BDD7EE"),
}
STATUS_FILL = {
    "PASS":    PatternFill("solid", fgColor="C6EFCE"),
    "FAIL":    PatternFill("solid", fgColor="FFC7CE"),
    "BLOCKED": PatternFill("solid", fgColor="FFF2CC"),
    "NOT RUN": PatternFill("solid", fgColor="BDD7EE"),
}
SEVERITY_FILL = {
    # PL (current)
    "Krytyczna": PatternFill("solid", fgColor="C00000"),
    "Wysoka":    PatternFill("solid", fgColor="F8CBAD"),
    "Średnia":   PatternFill("solid", fgColor="FFF2CC"),
    "Niska":     PatternFill("solid", fgColor="D9D9D9"),
    # Legacy English still recognized
    "Critical": PatternFill("solid", fgColor="C00000"),
    "High":     PatternFill("solid", fgColor="F8CBAD"),
    "Medium":   PatternFill("solid", fgColor="FFF2CC"),
    "Low":      PatternFill("solid", fgColor="D9D9D9"),
}
SEVERITY_FONT_WHITE = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
CATEGORY_FONT = Font(name="Calibri", size=10, bold=True)

# Column widths per sheet
CASES_WIDTHS = {
    # Polish (current)
    "ID": 11, "Tytuł": 42, "Priorytet": 14,
    "Warunki wstępne": 32, "Dane testowe": 30, "Kroki": 55, "Oczekiwany rezultat": 42,
    "Wynik": 12, "ID buga": 14, "Notatki": 35,
    # Legacy English still recognised
    "Title": 42, "Priority": 14, "Type": 12,
    "Preconditions": 32, "Test Data": 30, "Steps": 55, "Expected Result": 42,
    "Result": 12, "Bug ID": 14, "Notes": 35,
    "Suite": 20,
}
REPORT_WIDTHS = {
    # Polish (current)
    "ID": 11, "Status": 11, "Rzeczywisty rezultat": 60, "Waga": 12,
    "Zrzut ekranu": 26, "Reprodukcja": 55,
    # Legacy English still recognised
    "Actual Result": 60, "Severity": 12, "Screenshot": 26, "Repro": 55,
}
BUGS_WIDTHS = {
    "ID": 14, "Tytuł": 42, "Priorytet": 14, "Ważność/dotkliwość": 18,
    "Powiązany scenariusz testowy": 14,
    "Data zgłoszenia": 14,
    "URL": 55, "Przeglądarka": 20, "System operacyjny": 18, "Rozdzielczość": 14,
    "Kroki reprodukcji": 55,
    "Oczekiwany rezultat": 42, "Rzeczywisty rezultat": 42,
    "Screenshot": 26, "Wpływ na użytkownika": 45,
    # legacy still recognized
    "Ważność": 12, "Waga": 12,
}


def load_tsv(path: Path) -> list[list[str]]:
    return list(csv.reader(io.StringIO(path.read_text()), delimiter="\t"))


def load_meta(md_path: Path) -> dict:
    """Extract Feature / Priority / URL from the .md Brief section.
    Falls back to empty strings if missing."""
    meta = {"feature": "", "priority": "", "url": ""}
    if not md_path.exists():
        return meta
    text = md_path.read_text()
    # H1 like: "# QA run — Wyszukiwarka instytucji — 2026-08-13 (run 2)"
    h1 = re.search(r'^#\s+QA run\s+—\s+(.+?)\s+—\s+\d{4}-\d{2}-\d{2}', text, re.M)
    if h1:
        meta["feature"] = h1.group(1).strip()
    # Brief-section fields (authoritative if present)
    m = re.search(r'\*\*Feature:\*\*\s*(.+)', text)
    if m:
        meta["feature"] = m.group(1).strip()
    m = re.search(r'\*\*Priori(?:ty|tet):\*\*\s*(\S+)', text)
    if m:
        meta["priority"] = m.group(1).strip()
    m = re.search(r'\*\*URL:\*\*\s*(\S+)', text)
    if m:
        meta["url"] = m.group(1).strip()
    return meta


def title_text(meta: dict) -> str:
    parts = []
    if meta.get("feature"):
        parts.append(f"Testowana funkcjonalność: {meta['feature']}")
    if meta.get("priority"):
        parts.append(meta["priority"])
    return "  ·  ".join(parts)


def style_sheet(ws, rows: list[list[str]], widths: dict[str, int],
                category_fills: dict[int, dict], meta: dict):
    if not rows:
        return
    headers = rows[0]
    ncols = len(headers)
    origin = origin_of(meta.get("url", ""))

    # Column widths
    for idx, name in enumerate(headers, start=1):
        w = widths.get(name, 18)
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Row 1 — column headers
    for col_idx, val in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=val)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = BORDER
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # Body — start at row 2
    for offset, row in enumerate(rows[1:]):
        r_i = 2 + offset
        zebra = (offset % 2 == 1)  # alternate starting with unshaded first data row
        for c_i, val in enumerate(row, start=1):
            header = headers[c_i - 1]
            cell = ws.cell(row=r_i, column=c_i, value=val)
            cell.border = BORDER
            cell.font = BODY_FONT_BOLD if header in BOLD_COLUMNS else BODY_FONT

            # Category coloring
            colored = False
            for col_index, fill_map in category_fills.items():
                if c_i == col_index and val in fill_map:
                    cell.fill = fill_map[val]
                    if val == "Critical":
                        cell.font = SEVERITY_FONT_WHITE
                    elif header in CATEGORY_PLAIN_COLUMNS:
                        cell.font = BODY_FONT
                    else:
                        cell.font = CATEGORY_FONT
                    cell.alignment = BODY_ALIGN_CENTER
                    colored = True
                    break

            if not colored:
                if zebra:
                    cell.fill = ZEBRA_FILL
                if header == "ID" or (isinstance(val, str) and len(val) <= 3):
                    cell.alignment = BODY_ALIGN_CENTER
                else:
                    cell.alignment = BODY_ALIGN_TOP

                # Rich text: bold for "quoted" spans, blue+underline for URLs/paths.
                # URLs are visually styled but NOT set as cell hyperlink — Excel forces
                # the whole cell to link, which hurts UX (selecting text triggers navigation).
                # If you want a clickable link back, add `cell.hyperlink = href` here.
                if header in RICH_TEXT_COLUMNS and isinstance(val, str) and val and val != "-":
                    rt, _href = build_rich_text(val, origin)
                    cell.value = rt


RICH_TOKEN_RE = re.compile(
    r'"([^"]+)"'                                    # (1) "quoted" → bold
    r'|(https?://[^\s"<>)]+)'                       # (2) full URL → link
    r'|(/[a-zA-Z][\w\-/.]*(?:\?[^\s"<>)]+)?)'       # (3) relative path → link (needs base)
)


def build_rich_text(text: str, origin: str | None):
    """Parse text and return (CellRichText, first_href_or_none). If origin is given,
    relative paths starting with '/' are expanded to full URLs for the cell hyperlink."""
    rt = CellRichText()
    first_href = None
    last = 0
    for m in RICH_TOKEN_RE.finditer(text):
        if m.start() > last:
            rt.append(TextBlock(RT_PLAIN, text[last:m.start()]))
        quoted, url, path = m.group(1), m.group(2), m.group(3)
        if quoted is not None:
            # Preserve the quotes visually and bold the whole thing including them
            rt.append(TextBlock(RT_BOLD, f'"{quoted}"'))
        elif url is not None:
            rt.append(TextBlock(RT_LINK, url))
            if first_href is None:
                first_href = url
        elif path is not None:
            rt.append(TextBlock(RT_LINK, path))
            if first_href is None and origin:
                first_href = origin.rstrip("/") + path
        last = m.end()
    if last < len(text):
        rt.append(TextBlock(RT_PLAIN, text[last:]))
    return rt, first_href


def origin_of(url: str) -> str | None:
    if not url:
        return None
    p = urlparse(url)
    if not p.scheme or not p.netloc:
        return None
    return f"{p.scheme}://{p.netloc}"


def category_fills_for(headers: list[str], mapping: dict[str, dict]) -> dict[int, dict]:
    """Resolve header names → 1-based column indices for the fill maps."""
    out = {}
    for name, fill_map in mapping.items():
        if name in headers:
            out[headers.index(name) + 1] = fill_map
    return out


# Section header for per-bug tabs (row between title bar and field rows)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")  # very light blue-gray, subtle divider
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="305496")  # dark blue text
SECTION_ALIGN = Alignment(horizontal="left", vertical="center", indent=1)

# Label cell style for bug tabs (subtle gray, dark text — not shouty)
BUG_LABEL_FILL = PatternFill("solid", fgColor="F2F2F2")
BUG_LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="1F3864")
BUG_LABEL_ALIGN = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)

# Ordered sections mapping — each section groups a set of field headers
BUG_SECTIONS = [
    ("Metadane", ["ID", "Tytuł", "Priorytet", "Ważność/dotkliwość", "Powiązany scenariusz testowy", "Data zgłoszenia"]),
    ("Środowisko", ["URL", "Przeglądarka", "System operacyjny", "Rozdzielczość"]),
    ("Reprodukcja", ["Kroki reprodukcji"]),
    ("Rezultat", ["Oczekiwany rezultat", "Rzeczywisty rezultat"]),
    ("Analiza wpływu", ["Wpływ na użytkownika"]),
    ("Załączniki", ["Screenshot"]),
]


def add_bug_tabs(wb, bugs_rows: list[list[str]]):
    """Render each bug row as a separate vertical (label | value) tab with sections."""
    if len(bugs_rows) < 2:
        return
    headers = bugs_rows[0]
    id_idx = headers.index("ID") if "ID" in headers else 0
    # Prefer "Ważność/dotkliwość", fall back to legacy names
    for candidate in ("Ważność/dotkliwość", "Ważność", "Waga"):
        if candidate in headers:
            waga_idx = headers.index(candidate)
            break
    else:
        waga_idx = None
    title_idx = headers.index("Tytuł") if "Tytuł" in headers else 1

    # Tab color by severity — accepts PL + legacy EN values
    waga_colors = {
        "Krytyczna": "C00000", "Wysoka": "E97132", "Średnia": "FFC000", "Niska": "A6A6A6",
        "Critical":  "C00000", "High":   "E97132", "Medium":  "FFC000", "Low":   "A6A6A6",
    }

    for bug_row in bugs_rows[1:]:
        bug_id = bug_row[id_idx] if id_idx < len(bug_row) else "BUG"
        tab_name = re.sub(r'[\\/?*\[\]]', '-', bug_id)[:31]
        waga = bug_row[waga_idx] if waga_idx is not None and waga_idx < len(bug_row) else ""
        ws = wb.create_sheet(tab_name)
        ws.sheet_properties.tabColor = waga_colors.get(waga, "C00000")

        # Title bar (row 1)
        title = bug_row[title_idx] if title_idx < len(bug_row) else bug_id
        ws.cell(row=1, column=1, value=f"{bug_id} — {title}")
        ws.merge_cells("A1:B1")
        title_cell = ws.cell(row=1, column=1)
        title_cell.fill = TITLE_FILL
        title_cell.font = TITLE_FONT
        title_cell.alignment = TITLE_ALIGN
        ws.row_dimensions[1].height = 32

        # Vertical layout — 2 columns
        # A width = 32 to accommodate longest label "Powiązany scenariusz testowy" (28 chars) w 10pt
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 85
        # Ukryj wszystkie kolumny poza A:B — do samego końca arkusza Excel (XFD = 16384)
        # + thick frame wokół content (poniżej) daje bug tab wyglądający jak "karta"
        from openpyxl.worksheet.dimensions import ColumnDimension
        # Jeden ColumnDimension range zamiast 16k obiektów — RLE-compressed w XLSX
        hidden_dim = ColumnDimension(ws, min=3, max=16384, hidden=True)
        ws.column_dimensions['C'] = hidden_dim

        # Field lookup for random access
        field_values = {h: (bug_row[i] if i < len(bug_row) else "") for i, h in enumerate(headers)}
        # Track screenshot path & row (for embedding after row rendering completes)
        screenshot_row_after = None
        screenshot_path = field_values.get("Screenshot", "")

        r = 2  # current row cursor
        for section_name, section_fields in BUG_SECTIONS:
            # Section header row (merged, medium blue)
            sec_cell = ws.cell(row=r, column=1, value=section_name)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            sec_cell.fill = SECTION_FILL
            sec_cell.font = SECTION_FONT
            sec_cell.alignment = SECTION_ALIGN
            sec_cell.border = BORDER
            ws.row_dimensions[r].height = 24
            r += 1

            # Special: "Załączniki" section renders only embedded images (no label + no path)
            if section_name == "Załączniki":
                for field in section_fields:
                    value = field_values.get(field, "")
                    if not value or value == "-":
                        continue
                    img_path = Path(value)
                    if not img_path.exists():
                        continue
                    try:
                        from PIL import Image as PILImage
                        with PILImage.open(img_path) as im:
                            orig_w, orig_h = im.size
                        max_w = 650
                        if orig_w > max_w:
                            w = max_w
                            h = int(orig_h * max_w / orig_w)
                        else:
                            w, h = orig_w, orig_h
                    except Exception:
                        w, h = 650, 400
                    xl_img = XLImage(str(img_path))
                    xl_img.width = w
                    xl_img.height = h
                    xl_img.anchor = f'A{r}'
                    ws.add_image(xl_img)
                    rows_needed = max(1, (h + 10) // 15)
                    for i in range(rows_needed):
                        ws.row_dimensions[r + i].height = 15
                    r += rows_needed
                continue  # skip normal field loop for Załączniki

            # Field rows within section — always render label|value
            # (skip label ONLY gdy field name == section name, np. Środowisko section z Środowisko field)
            for field in section_fields:
                if field not in field_values:
                    continue
                value = field_values[field]
                skip_label = (field == section_name)

                if skip_label:
                    val_cell = ws.cell(row=r, column=1, value=value)
                    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
                    val_cell.border = BORDER
                    val_cell.font = BODY_FONT
                    val_cell.alignment = Alignment(horizontal="left", vertical="center", indent=3, wrap_text=True)
                else:
                    label_cell = ws.cell(row=r, column=1, value=field)
                    label_cell.fill = BUG_LABEL_FILL
                    label_cell.font = BUG_LABEL_FONT
                    label_cell.alignment = BUG_LABEL_ALIGN
                    label_cell.border = BORDER

                    val_cell = ws.cell(row=r, column=2, value=value)
                    val_cell.border = BORDER
                    val_cell.font = BODY_FONT
                    val_cell.alignment = BODY_ALIGN_TOP

                # Severity (Ważność/dotkliwość / Ważność / Waga) row color
                if field in ("Ważność/dotkliwość", "Ważność", "Waga") and value in SEVERITY_FILL:
                    val_cell.fill = SEVERITY_FILL[value]
                    if value in ("Critical", "Krytyczna"):
                        val_cell.font = SEVERITY_FONT_WHITE
                    else:
                        val_cell.font = CATEGORY_FONT
                    val_cell.alignment = BODY_ALIGN_CENTER
                # Priorytet row color (case priority palette)
                elif field == "Priorytet" and value in PRIORITY_FILL:
                    val_cell.fill = PRIORITY_FILL[value]
                    val_cell.font = BODY_FONT_BOLD
                    val_cell.alignment = BODY_ALIGN_CENTER
                # Oczekiwany rezultat — jasnozielone tło (co powinno być)
                elif field == "Oczekiwany rezultat" and value and value != "-":
                    val_cell.fill = PatternFill("solid", fgColor="E2EFDA")
                # Rzeczywisty rezultat — jasnoczerwone tło (co jest = bug)
                elif field == "Rzeczywisty rezultat" and value and value != "-":
                    val_cell.fill = PatternFill("solid", fgColor="FCE4E4")


                # Rich text
                if field in RICH_TEXT_COLUMNS and isinstance(value, str) and value and value != "-":
                    rt, _href = build_rich_text(value, None)
                    val_cell.value = rt

                # Row height: min 22pt; higher for long values
                # Merged row (skip_label) has wider effective width
                chars_per_line = 100 if skip_label else 70
                base_lines = 1
                if isinstance(value, str) and value:
                    base_lines = value.count('\n') + 1 + max(0, (len(value) - 1) // chars_per_line)
                label_lines = 1 if skip_label else max(1, (len(field) - 1) // 30 + 1)
                lines = max(base_lines, label_lines)
                extra = 1 if skip_label else 0
                ws.row_dimensions[r].height = max(22, min(240, (lines + extra) * 16))

                r += 1

        ws.sheet_view.showGridLines = False


LINK_FONT_BOLD = Font(name="Calibri", size=10, bold=True, color="0563C1", underline="single")
LINK_FONT_PLAIN = Font(name="Calibri", size=10, color="0563C1", underline="single")


def add_bug_id_hyperlinks(wb):
    """Bidirectional internal links via Hyperlink object z location=
    (proper XML: <hyperlink ref location display>, bez TargetMode=External).
    Działa w Excelu, Numbers, i Google Sheets."""
    # Test cases → bug tab
    if "Test cases" in wb.sheetnames:
        ws = wb["Test cases"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        for col_name in ("ID buga", "Bug ID"):
            if col_name in headers:
                col_idx = headers.index(col_name) + 1
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r, column=col_idx)
                    val = cell.value
                    if isinstance(val, str) and val and val != "-" and val in wb.sheetnames:
                        cell.hyperlink = Hyperlink(
                            ref=cell.coordinate,
                            location=f"'{val}'!A1",
                            display=val,
                        )
                        cell.font = LINK_FONT_BOLD
                break

    # Bug tab → Test cases row
    case_id_row_map = {}
    if "Test cases" in wb.sheetnames:
        tc_ws = wb["Test cases"]
        tc_headers = [tc_ws.cell(row=1, column=c).value for c in range(1, tc_ws.max_column + 1)]
        if "ID" in tc_headers:
            id_col = tc_headers.index("ID") + 1
            for r in range(2, tc_ws.max_row + 1):
                v = tc_ws.cell(row=r, column=id_col).value
                if isinstance(v, str) and v:
                    case_id_row_map[v] = r

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith(("INST-BR-", "BR-")) and "-BR-" not in sheet_name:
            continue
        ws = wb[sheet_name]
        for r in range(1, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            if label == "Powiązany scenariusz testowy":
                val_cell = ws.cell(row=r, column=2)
                case_id = val_cell.value
                if isinstance(case_id, str) and case_id in case_id_row_map:
                    target_row = case_id_row_map[case_id]
                    val_cell.hyperlink = Hyperlink(
                        ref=val_cell.coordinate,
                        location=f"'Test cases'!A{target_row}",
                        display=case_id,
                    )
                    val_cell.font = LINK_FONT_PLAIN
                break


def build(base: Path):
    cases_rows = load_tsv(base.with_name(base.name + "-cases.tsv"))
    report_rows = load_tsv(base.with_name(base.name + "-report.tsv"))
    bugs_path = base.with_name(base.name + "-bugs.tsv")
    bugs_rows = load_tsv(bugs_path) if bugs_path.exists() else None
    meta = load_meta(base.with_suffix(".md")) if base.suffix else load_meta(base.with_name(base.name + ".md"))

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Test cases"
    ws1.sheet_properties.tabColor = "305496"
    style_sheet(
        ws1, cases_rows, CASES_WIDTHS,
        category_fills_for(cases_rows[0], {
            "Priorytet": PRIORITY_FILL, "Priority": PRIORITY_FILL,
            "Wynik": STATUS_FILL, "Result": STATUS_FILL,
        }),
        meta,
    )

    ws2 = wb.create_sheet("Execution report")
    ws2.sheet_properties.tabColor = "305496"
    style_sheet(
        ws2, report_rows, REPORT_WIDTHS,
        category_fills_for(report_rows[0], {
            "Status": STATUS_FILL,
            "Waga": SEVERITY_FILL, "Severity": SEVERITY_FILL,
        }),
        meta,
    )

    if bugs_rows:
        add_bug_tabs(wb, bugs_rows)

    # Post-processing: dodaj bidirectional hyperlinki między Test cases i bug tabs
    add_bug_id_hyperlinks(wb)

    out = base.with_suffix(".xlsx") if base.suffix else base.with_name(base.name + ".xlsx")
    wb.save(out)
    print(f"Wrote {out}" + (f" (with Bug tracker: {len(bugs_rows)-1} bugs)" if bugs_rows else ""))


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(__doc__, file=sys.stderr)
        sys.exit(1)
    build(Path(sys.argv[1]))
